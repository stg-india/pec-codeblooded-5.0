import csv
from openpyxl import load_workbook

wb = load_workbook(filename='H-1B_Disclosure_Data_FY16.xlsx')
sheet = wb.active

csv_data=[]
for value in sheet.iter_rows(values_only=True):
    csv_data.append(list(value))

with open('H-1B_Disclosure_Data_FY16.csv','w') as csv_obj:
    writer = csv.writer(csv_obj,delimiter=',')
    for line in csv_data:
        writer.writerow(line)